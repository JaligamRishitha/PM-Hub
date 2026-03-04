import io
from datetime import date

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.security import get_current_user
from app.models.activity import ProjectActivity
from app.models.cbs import CBS
from app.models.document import Document
from app.models.issue import Issue
from app.models.milestone import Milestone
from app.models.project import Project
from app.models.risk import Risk
from app.models.safety_incident import SafetyIncident
from app.models.user import User
from app.models.variation import ApprovedVariation

router = APIRouter(prefix="/api/export", tags=["export"])


def _to_str(val):
    if val is None:
        return ""
    if isinstance(val, date):
        return val.isoformat()
    return str(val)


@router.get("/projects")
def export_projects(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    wb = Workbook()

    # Projects sheet
    ws = wb.active
    ws.title = "Projects"
    ws.append(["ID", "Name", "Client", "Start Date", "End Date", "Status", "Budget", "LD Rate"])
    for p in db.query(Project).all():
        ws.append([
            p.id, p.name, p.client,
            _to_str(p.start_date), _to_str(p.end_date),
            p.status, float(p.total_budget), float(p.daily_ld_rate),
        ])

    # Activities sheet
    ws2 = wb.create_sheet("Activities")
    ws2.append([
        "ID", "Project ID", "Activity", "Phase",
        "Planned Start", "Planned Finish", "Actual Start", "Actual Finish",
        "Completion %", "Status", "Delay Days",
    ])
    for a in db.query(ProjectActivity).all():
        ws2.append([
            a.id, a.project_id, a.activity_name, a.phase,
            _to_str(a.planned_start), _to_str(a.planned_finish),
            _to_str(a.actual_start), _to_str(a.actual_finish),
            float(a.completion_pct), a.status, a.delay_days,
        ])

    # Milestones sheet
    ws3 = wb.create_sheet("Milestones")
    ws3.append(["ID", "Project ID", "Name", "Phase", "Planned Date", "Actual Date", "Critical", "Status", "Delay Days"])
    for m in db.query(Milestone).all():
        ws3.append([
            m.id, m.project_id, m.name, m.phase,
            _to_str(m.planned_date), _to_str(m.actual_date),
            m.is_critical, m.status, m.delay_days,
        ])

    # CBS sheet
    ws4 = wb.create_sheet("CBS")
    ws4.append(["ID", "Project ID", "WBS Code", "Description", "Budget", "Actual", "Variation", "Variance"])
    for c in db.query(CBS).all():
        ws4.append([
            c.id, c.project_id, c.wbs_code, c.description,
            float(c.budget_cost), float(c.actual_cost),
            float(c.approved_variation), c.variance,
        ])

    # Safety sheet
    ws5 = wb.create_sheet("Safety Incidents")
    ws5.append(["ID", "Project ID", "Type", "Severity", "Reported", "Resolved", "Status", "Penalty", "Lost Time Hours", "Location"])
    for s in db.query(SafetyIncident).all():
        ws5.append([
            s.id, s.project_id, s.incident_type, s.severity,
            _to_str(s.reported_date), _to_str(s.resolved_date),
            s.status, float(s.penalty_cost),
            float(s.lost_time_hours or 0), s.location,
        ])

    # Risks sheet
    ws6 = wb.create_sheet("Risks")
    ws6.append(["ID", "Project ID", "Title", "Category", "Probability", "Impact", "Score", "Status", "Owner", "Cost Exposure", "Mitigation"])
    for r in db.query(Risk).all():
        ws6.append([
            r.id, r.project_id, r.title, r.category,
            r.probability, r.impact, r.risk_score,
            r.status, r.owner, float(r.cost_exposure or 0),
            r.mitigation_plan,
        ])

    # Issues sheet
    ws7 = wb.create_sheet("Issues")
    ws7.append(["ID", "Project ID", "Title", "Priority", "Status", "Assigned To", "Due Date", "Raised Date"])
    for i in db.query(Issue).all():
        ws7.append([
            i.id, i.project_id, i.title, i.priority,
            i.status, i.assigned_to, _to_str(i.due_date),
            _to_str(i.raised_date),
        ])

    # Variations sheet
    ws8 = wb.create_sheet("Variations")
    ws8.append(["ID", "Project ID", "Variation Code", "Description", "Value", "Cost Impact", "Schedule Impact", "Status", "Approved Date"])
    for v in db.query(ApprovedVariation).all():
        ws8.append([
            v.id, v.project_id, v.variation_code, v.description,
            float(v.value or 0), float(v.cost_impact or 0),
            v.schedule_impact_days, v.approval_status,
            _to_str(v.approval_date),
        ])

    # Documents sheet
    ws9 = wb.create_sheet("Documents")
    ws9.append(["ID", "Project ID", "Doc Code", "Title", "Category", "Revision", "Status", "Uploaded By"])
    for d in db.query(Document).all():
        ws9.append([
            d.id, d.project_id, d.doc_code, d.title,
            d.category, d.revision, d.status, d.uploaded_by,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pm_hub_export.xlsx"},
    )
