"""Seed script to populate the database with PM user and Primavera P6 Excel data."""
import os
import re
import sys
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

from app.core.database import Base, engine, SessionLocal
from app.core.security import hash_password
from app.models.user import User
from app.models.organization import Organization
from app.models.programme import Programme
from app.models.project import Project
from app.models.activity import ProjectActivity
from app.models.milestone import Milestone
from app.models.cbs import CBS
from app.models.wbs import WBSItem
from app.models.risk import Risk
from app.models.safety_incident import SafetyIncident

# Import all models so tables are created
import app.models  # noqa: F401

# ── Excel files (each file = one project) ────────────────────────
EXCEL_DIR = "/data"  # mounted volume in Docker

EXCEL_FILES = [
    "Primavera P6 - Activities of __Awaiting Gate D__ EJH (NTR) A47_A11 Thickthorn Jtn, NR4 6TD - 33kV Diversion.xlsx",
    "Primavera P6 - Activities of Awaiting Gate D - Kidbrooke Park Road SE3 9PX.xlsx",
    "Primavera P6 - Activities of Eastfield Est, Acacia Road, Mitcham, CR4 1BX, Green Recovery - EPN EV Fuel Stations (Howe G.xlsx",
    "Primavera P6 - Activities of EDL Harlow Science Park, Zone 5 London Road, CM17 9NA, EZW (ANM) Hulcote Farm, Salford Road.xlsx",
    "Primavera P6 - Activities of EDL Harlow Science Park, Zone 5 London Road, CM17 9NA, EZW (ANM) Hulcote Farm, Salford Road (1).xlsx",
    "Primavera P6 - Activities of EJHo Boscombe Road LU5 4LT.xlsx",
    "Primavera P6 - Activities of EZW (ANM) Hulcote Farm, Salford Road, MK17 8BS.xlsx",
    "Primavera P6 - Activities of Green Recovery - EPN EV Fuel Stations (Howe Green).xlsx",
    "Primavera P6 - Activities of Hackney Grid 66kV Switchgear Replacement.xlsx",
    "Primavera P6 - Activities of RAYLEIGH_CHELMSFORD(PSA).xlsx",
]

# ── Manual cost overrides per project (keyword in name → costs) ──
# Keys are substrings matched against the project name (case-insensitive)
COST_OVERRIDES = {
    "edl harlow science park": {
        "budget": 345666.00,   # BL Project Total Cost
        "actual": 300600.00,   # Actual Total Cost
    },
}


def extract_project_name(filename):
    """Extract project name from the filename pattern."""
    name = filename.replace(".xlsx", "")
    name = re.sub(r"^Primavera P6 - Activities of\s*", "", name)
    # Clean up double underscores used as escaping
    name = name.replace("__", "")
    name = name.strip()
    return name


def to_date(value):
    """Convert Excel cell value to a Python date or None."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Handle string dates like '01-Jul-22 09:00 AM'
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in (
            "%d-%b-%y %I:%M %p",   # 01-Jul-22 09:00 AM
            "%d-%b-%Y %I:%M %p",   # 01-Jul-2022 09:00 AM
            "%d-%b-%y",             # 01-Jul-22
            "%d-%b-%Y",             # 01-Jul-2022
            "%Y-%m-%d",             # 2022-07-01
            "%d/%m/%Y",             # 01/07/2022
            "%m/%d/%Y",             # 07/01/2022
        ):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def to_float(value, default=0.0):
    """Convert Excel cell value to float."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def to_pct(value):
    """Convert completion percentage (could be 0-1 or 0-100) to 0-100 scale."""
    v = to_float(value, 0.0)
    if 0 < v <= 1.0:
        return round(v * 100, 2)
    return round(v, 2)


def determine_phase(activity_name, completion_pct):
    """Fallback: derive a phase from the activity name or completion when
    WBS section tracking does not yield a gate-based phase."""
    name_lower = (activity_name or "").lower()
    if any(kw in name_lower for kw in ["design", "survey", "assess", "plan"]):
        return "Gate B-C"
    if any(kw in name_lower for kw in ["commission", "handover", "energi", "go-live", "close"]):
        return "Closure"
    if any(kw in name_lower for kw in ["test", "inspect", "check", "qc", "qa"]):
        return "Gate C-D"
    return "Gate C-D"


def detect_section_phase(activity_name):
    """Detect if a WBS summary row marks a gate section boundary.
    Returns the new phase name or None if not a boundary."""
    name_lower = (activity_name or "").lower().strip()
    if name_lower in ("gate b-c", "gate b - c"):
        return "Gate B-C"
    if name_lower in ("gate c-d", "gate c - d"):
        return "Gate C-D"
    if "pre gate b" in name_lower or "pre-gate b" in name_lower:
        return "Pre Gate B"
    if name_lower == "closure" or name_lower == "project closeout milestones":
        return "Closure"
    return None


def seed_risks(db, project_id, project_name, project_budget):
    """Seed realistic risk register entries for a project."""
    import random
    random.seed(hash(project_name))  # Deterministic per project

    budget = float(project_budget or 100000)

    risk_templates = [
        {"title": "Supply chain delay", "category": "Schedule", "prob": 3, "impact": 4,
         "desc": "Key materials or equipment may face supply chain delays",
         "mitigation": "Maintain buffer stock and identify alternative suppliers",
         "exposure_pct": 0.05},
        {"title": "Ground conditions uncertainty", "category": "Technical", "prob": 2, "impact": 3,
         "desc": "Unexpected ground conditions during excavation",
         "mitigation": "Conduct additional ground surveys before excavation",
         "exposure_pct": 0.03},
        {"title": "Permit / consent delays", "category": "Regulatory", "prob": 3, "impact": 3,
         "desc": "Planning permissions or regulatory consents may be delayed",
         "mitigation": "Early engagement with local authority and stakeholders",
         "exposure_pct": 0.02},
        {"title": "Resource availability", "category": "Resource", "prob": 2, "impact": 2,
         "desc": "Specialist labour may not be available when needed",
         "mitigation": "Advance booking of specialist contractors",
         "exposure_pct": 0.04},
        {"title": "Weather impact on programme", "category": "Environmental", "prob": 3, "impact": 2,
         "desc": "Adverse weather conditions may delay outdoor works",
         "mitigation": "Build weather contingency into programme float",
         "exposure_pct": 0.02},
        {"title": "Third party interface risk", "category": "External", "prob": 2, "impact": 4,
         "desc": "Dependencies on third party utilities or landowners",
         "mitigation": "Early engagement and formal agreements with third parties",
         "exposure_pct": 0.06},
        {"title": "Design change during construction", "category": "Technical", "prob": 2, "impact": 3,
         "desc": "Late design changes requiring rework",
         "mitigation": "Design freeze checkpoint before construction start",
         "exposure_pct": 0.04},
        {"title": "Health & safety incident", "category": "Safety", "prob": 1, "impact": 5,
         "desc": "Potential for workplace injury or safety event",
         "mitigation": "Robust RAMS, toolbox talks, and safety audits",
         "exposure_pct": 0.01},
    ]

    # Each project gets 3-6 risks
    num_risks = random.randint(3, 6)
    selected = random.sample(risk_templates, min(num_risks, len(risk_templates)))

    for i, tmpl in enumerate(selected):
        # Some risks are closed/mitigated
        status = random.choice(["Open", "Open", "Open", "Mitigated", "Closed"])
        exposure = round(budget * tmpl["exposure_pct"], 2)

        risk = Risk(
            project_id=project_id,
            risk_code=f"R-{project_id:03d}-{i+1:02d}",
            title=tmpl["title"],
            description=tmpl["desc"],
            category=tmpl["category"],
            probability=tmpl["prob"],
            impact=tmpl["impact"],
            mitigation_plan=tmpl["mitigation"],
            owner="Alex Johnson",
            status=status,
            cost_exposure=exposure,
            identified_date=date(2024, 6, 1) + timedelta(days=random.randint(0, 300)),
        )
        db.add(risk)

    db.flush()
    open_count = sum(1 for r in selected if True)  # We'll count after commit
    return num_risks


def seed_safety_incidents(db, project_id, project_name, project_start):
    """Seed realistic safety incident records for a project."""
    import random
    random.seed(hash(project_name) + 42)  # Deterministic but different from risks

    incident_templates = [
        {"type": "Near Miss", "severity": "Minor", "desc": "Unsecured cable tray near walkway"},
        {"type": "Observation", "severity": "Minor", "desc": "Incomplete barrier around excavation"},
        {"type": "Near Miss", "severity": "Moderate", "desc": "Vehicle reversing near pedestrian route without banksman"},
        {"type": "Observation", "severity": "Minor", "desc": "PPE not worn in designated area"},
        {"type": "Injury", "severity": "Minor", "desc": "Minor hand laceration during cable pulling"},
        {"type": "Environmental", "severity": "Minor", "desc": "Small fuel spill from generator contained on site"},
        {"type": "Observation", "severity": "Minor", "desc": "Fire extinguisher past inspection date"},
        {"type": "Near Miss", "severity": "Moderate", "desc": "Dropped tool from height — exclusion zone in place"},
        {"type": "Injury", "severity": "Moderate", "desc": "Slip on wet surface resulting in twisted ankle"},
        {"type": "Observation", "severity": "Minor", "desc": "Housekeeping improvement needed in switch room"},
    ]

    # Each project gets 2-5 incidents
    num_incidents = random.randint(2, 5)
    selected = random.sample(incident_templates, min(num_incidents, len(incident_templates)))
    start = project_start or date(2024, 6, 1)

    for tmpl in selected:
        reported = start + timedelta(days=random.randint(30, 500))
        # Most incidents are resolved
        is_resolved = random.random() < 0.75
        resolved = reported + timedelta(days=random.randint(1, 14)) if is_resolved else None
        status = random.choice(["Resolved", "Closed"]) if is_resolved else random.choice(["Open", "Under Investigation"])

        incident = SafetyIncident(
            project_id=project_id,
            incident_type=tmpl["type"],
            severity=tmpl["severity"],
            reported_date=reported,
            resolved_date=resolved,
            status=status,
            description=tmpl["desc"],
            lost_time_hours=random.choice([0, 0, 0, 2, 4, 8]) if tmpl["type"] == "Injury" else 0,
            location=project_name[:100],
            reported_by="Site Supervisor",
        )
        db.add(incident)

    db.flush()
    return num_incidents


def import_excel(db, filepath, filename, programme_id=None):
    """Import one Primavera P6 Excel file as a project + activities."""
    project_name = extract_project_name(filename)
    print(f"  Importing: {project_name}")

    wb = load_workbook(filepath, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if len(rows) < 2:
        print(f"    Skipped (no data rows): {filename}")
        return

    # Header row (row 0) — data rows start from row 1
    header = rows[0]
    data_rows = rows[1:]

    # Extract project-level costs from the FIRST data row (project summary)
    # The first row contains the total project budget/actual/forecast
    proj_summary_row = data_rows[0] if data_rows else None
    project_budget = to_float(proj_summary_row[3] if proj_summary_row and len(proj_summary_row) > 3 else None)
    project_actual = to_float(proj_summary_row[4] if proj_summary_row and len(proj_summary_row) > 4 else None)
    project_forecast = to_float(proj_summary_row[6] if proj_summary_row and len(proj_summary_row) > 6 else None)

    # Apply manual cost overrides if defined
    for key, overrides in COST_OVERRIDES.items():
        if key in project_name.lower():
            project_budget = overrides.get("budget", project_budget)
            project_actual = overrides.get("actual", project_actual)
            project_forecast = overrides.get("forecast", project_forecast)
            print(f"    Applied cost overrides: budget={project_budget}, actual={project_actual}")
            break

    # Collect dates from all activities to derive project-level fields
    all_planned_starts = []
    all_planned_finishes = []

    activity_records = []
    wbs_records = []        # Collect WBS summary rows in order
    current_section_phase = "Pre Gate B"  # Track gate section from WBS headers
    current_wbs_sort = None  # Track which WBS node (by sort_order) activities belong to
    prev_row_was_wbs = False  # Track consecutive WBS rows

    for row in data_rows:
        # Skip completely empty rows
        if not row or (not row[0] and not row[1]):
            continue

        activity_code = str(row[0]).strip() if row[0] else None
        activity_name = str(row[1]).strip() if row[1] else None

        if not activity_name:
            continue

        # Check if this row is a WBS section header that defines a gate phase
        section = detect_section_phase(activity_name)
        if section:
            current_section_phase = section

        # Column indices (0-based): D=3, E=4, F=5, G=6, H=7, I=8, J=9, K=10, L=11, M=12, N=13, O=14, P=15, Q=16
        performance_pct = to_pct(row[8] if len(row) > 8 else None)  # Performance % Complete
        planned_start = to_date(row[9] if len(row) > 9 else None)
        planned_finish = to_date(row[10] if len(row) > 10 else None)
        actual_start = to_date(row[11] if len(row) > 11 else None)
        actual_finish = to_date(row[12] if len(row) > 12 else None)
        activity_pct = to_pct(row[16] if len(row) > 16 else None)   # Activity % Complete
        # Use Activity % Complete if available, fallback to Performance % Complete
        completion_pct = activity_pct if activity_pct > 0 else performance_pct

        # WBS summary rows (numeric-only codes like 1, 6, 31, 311, 441)
        # These are P6 WBS grouping nodes — collect them for WBS structure
        if activity_code and activity_code.isdigit():
            wbs_records.append({
                "code": activity_code,
                "name": activity_name,
                "sort_order": len(wbs_records),
                "consecutive": prev_row_was_wbs,  # True if previous row was also WBS
            })
            current_wbs_sort = len(wbs_records) - 1
            prev_row_was_wbs = True
            continue

        # Skip rows without any date info (likely summary/header rows)
        if not planned_start and not planned_finish and not actual_start:
            prev_row_was_wbs = False
            continue

        if planned_start:
            all_planned_starts.append(planned_start)
        if planned_finish:
            all_planned_finishes.append(planned_finish)

        # Use gate section phase from WBS headers, fallback to keyword detection
        phase = current_section_phase or determine_phase(activity_name, completion_pct)

        # Use today as fallback for required date fields
        p_start = planned_start or actual_start or date.today()
        p_finish = planned_finish or actual_finish or p_start

        activity_records.append({
            "activity_code": activity_code,
            "activity_name": activity_name,
            "phase": phase,
            "planned_start": p_start,
            "planned_finish": p_finish,
            "actual_start": actual_start,
            "actual_finish": actual_finish,
            "completion_pct": completion_pct,
            "is_milestone": (p_start == p_finish),
            "is_critical": False,
            "parent_wbs_sort": current_wbs_sort,
        })
        prev_row_was_wbs = False

    if not activity_records:
        print(f"    Skipped (no valid activities): {filename}")
        return

    # Derive project-level dates
    project_start = min(all_planned_starts) if all_planned_starts else date.today()
    project_end = max(all_planned_finishes) if all_planned_finishes else date.today()

    # Generate a short project code from the name
    words = project_name.split()
    code = "".join(w[0].upper() for w in words if w[0].isalpha())[:8]
    if not code:
        code = "PRJ"

    # ── Derive project phase from gate activities ─────────
    # Look for gate-related activities and their completion to determine phase
    gate_order = ["gate a", "gate b", "gate c", "design completion", "construction", "commissioned"]
    # Map gate activity names to dashboard phases
    gate_phase_map = {
        "gate a": "Gate B",        # Gate A done → awaiting Gate B
        "gate b": "Gate C",        # Gate B done → awaiting Gate C
        "gate c": "Design Completion",  # Gate C done → in Design Completion
        "design completion": "Construction",
        "construction": "Commissioned",
    }
    last_completed_gate = None
    for rec in activity_records:
        name_lower = (rec["activity_name"] or "").lower()
        pct = rec["completion_pct"]
        for gate_key in gate_order:
            if gate_key in name_lower and pct >= 100:
                last_completed_gate = gate_key

    if last_completed_gate and last_completed_gate in gate_phase_map:
        project_phase = gate_phase_map[last_completed_gate]
    else:
        # Fallback: derive from overall completion
        total_pct = sum(r["completion_pct"] for r in activity_records)
        avg_pct = total_pct / len(activity_records) if activity_records else 0
        if avg_pct >= 95:
            project_phase = "Commissioned"
        elif avg_pct >= 60:
            project_phase = "Construction"
        elif avg_pct >= 30:
            project_phase = "Design Completion"
        elif avg_pct >= 10:
            project_phase = "Gate C"
        else:
            project_phase = "Gate B"

    # Create project
    project = Project(
        programme_id=programme_id,
        name=project_name,
        code=code,
        client="UKPN",
        description=f"Primavera P6 imported project: {project_name}",
        start_date=project_start,
        end_date=project_end,
        contract_completion_date=project_end,
        status="Active",
        phase=project_phase,
        total_budget=round(project_budget, 2),
        contract_value=round(project_budget, 2),
        forecast_cost=round(project_forecast, 2),
        location="UK",
        project_manager="Alex Johnson",
    )
    db.add(project)
    db.flush()

    # ── Create WBS items from P6 WBS summary rows ──
    # Build hierarchy: known gate sections are top-level, consecutive WBS = deeper nesting
    TOP_LEVEL_NAMES = {
        "project summary", "pre gate b", "gate b-c", "gate b - c",
        "gate c-d", "gate c - d", "closure",
    }
    wbs_code_to_id = {}
    current_top_wbs = None    # Level 0 gate section
    current_child_wbs = None  # Level 1 sub-section

    for wrec in wbs_records:
        wname_lower = wrec["name"].lower().strip()
        is_top = wname_lower in TOP_LEVEL_NAMES

        if is_top:
            # Top-level gate section (level 0)
            wbs_item = WBSItem(
                project_id=project.id,
                parent_id=None,
                code=wrec["code"],
                name=wrec["name"],
                level=0,
                sort_order=wrec["sort_order"],
            )
            db.add(wbs_item)
            db.flush()
            current_top_wbs = wbs_item
            current_child_wbs = None
        elif wrec["consecutive"] and current_child_wbs:
            # Consecutive WBS right after another child → grandchild (level 2)
            wbs_item = WBSItem(
                project_id=project.id,
                parent_id=current_child_wbs.id,
                code=wrec["code"],
                name=wrec["name"],
                level=2,
                sort_order=wrec["sort_order"],
            )
            db.add(wbs_item)
            db.flush()
        else:
            # Child of current gate section (level 1)
            parent = current_top_wbs
            wbs_item = WBSItem(
                project_id=project.id,
                parent_id=parent.id if parent else None,
                code=wrec["code"],
                name=wrec["name"],
                level=1 if parent else 0,
                sort_order=wrec["sort_order"],
            )
            db.add(wbs_item)
            db.flush()
            current_child_wbs = wbs_item

        wbs_code_to_id[wrec["code"] + ":" + str(wrec["sort_order"])] = wbs_item.id

    # Build lookup: sort_order → WBSItem.id
    wbs_sort_to_id = {}
    for wrec in wbs_records:
        key = wrec["code"] + ":" + str(wrec["sort_order"])
        if key in wbs_code_to_id:
            wbs_sort_to_id[wrec["sort_order"]] = wbs_code_to_id[key]

    # Create activities
    for rec in activity_records:
        wbs_id = wbs_sort_to_id.get(rec.get("parent_wbs_sort"))

        activity = ProjectActivity(
            project_id=project.id,
            wbs_id=wbs_id,
            activity_code=rec["activity_code"],
            activity_name=rec["activity_name"],
            phase=rec["phase"],
            planned_start=rec["planned_start"],
            planned_finish=rec["planned_finish"],
            actual_start=rec["actual_start"],
            actual_finish=rec["actual_finish"],
            completion_pct=rec["completion_pct"],
            is_milestone=rec["is_milestone"],
            is_critical=rec["is_critical"],
        )
        db.add(activity)

    # ── Create milestones from milestone-like activities ──
    milestone_count = 0
    for rec in activity_records:
        name_lower = (rec["activity_name"] or "").lower()
        is_zero_duration = rec["planned_start"] == rec["planned_finish"]
        is_gate = any(kw in name_lower for kw in ["gate ", "milestone", "sign-off", "approval",
                       "handover", "energis", "go-live", "completion", "commissioning",
                       "start on site", "award", "consent"])
        if is_zero_duration or is_gate:
            milestone = Milestone(
                project_id=project.id,
                name=rec["activity_name"],
                phase=rec["phase"],
                planned_date=rec["planned_start"],
                actual_date=rec["actual_finish"],
                is_critical=is_gate,
            )
            db.add(milestone)
            milestone_count += 1

    # ── Create CBS record from project summary row costs ─
    if project_budget > 0 or project_actual > 0:
        cbs_entry = CBS(
            project_id=project.id,
            wbs_code="1.0",
            description=f"{project_name} - Total Cost",
            budget_cost=round(project_budget, 2),
            actual_cost=round(project_actual, 2),
            forecast_cost=round(project_forecast, 2),
        )
        db.add(cbs_entry)

    # ── Seed risk register and safety incidents ─────────
    risk_count = seed_risks(db, project.id, project_name, project_budget)
    incident_count = seed_safety_incidents(db, project.id, project_name, project_start)

    db.flush()
    print(f"    Created project (id={project.id}, phase={project_phase}) with {len(activity_records)} activities, {milestone_count} milestones, {len(wbs_records)} WBS items, {risk_count} risks, {incident_count} incidents")


def update_excel_data():
    """Re-import Excel data into the database, replacing existing project data."""
    Base.metadata.create_all(bind=engine)

    # Migrate: add phase column if missing
    from sqlalchemy import text
    with engine.connect() as conn:
        try:
            conn.execute(text(
                "ALTER TABLE projects ADD COLUMN IF NOT EXISTS phase VARCHAR(50) DEFAULT 'Gate B'"
            ))
            conn.commit()
        except Exception:
            conn.rollback()

    db = SessionLocal()
    try:
        # Find the programme to attach projects to
        programme = db.query(Programme).first()
        if not programme:
            print("No programme found. Run initial seed first.")
            sys.exit(1)

        programme_id = programme.id

        # Delete existing project data (cascade will handle activities, milestones, CBS)
        existing_projects = db.query(Project).all()
        if existing_projects:
            print(f"Removing {len(existing_projects)} existing projects and related data...")
            for proj in existing_projects:
                # Delete related records first
                db.query(ProjectActivity).filter(ProjectActivity.project_id == proj.id).delete()
                db.query(Milestone).filter(Milestone.project_id == proj.id).delete()
                db.query(CBS).filter(CBS.project_id == proj.id).delete()
                db.query(Risk).filter(Risk.project_id == proj.id).delete()
                db.query(SafetyIncident).filter(SafetyIncident.project_id == proj.id).delete()
                db.query(WBSItem).filter(WBSItem.project_id == proj.id).delete()
                db.delete(proj)
            db.flush()
            print("Existing project data removed.")

        # Re-import from Excel files
        print("Importing Primavera P6 Excel files...")
        for filename in EXCEL_FILES:
            filepath = os.path.join(EXCEL_DIR, filename)
            if os.path.exists(filepath):
                try:
                    import_excel(db, filepath, filename, programme_id=programme_id)
                except Exception as e:
                    print(f"    ERROR importing {filename}: {e}")
            else:
                print(f"    File not found: {filepath}")

        db.commit()
        print("Excel data updated successfully!")
    except Exception as e:
        db.rollback()
        print(f"Update failed: {e}")
        sys.exit(1)
    finally:
        db.close()


def seed():
    Base.metadata.create_all(bind=engine)

    # Migrate: add phase column if missing
    from sqlalchemy import text
    with engine.connect() as conn:
        try:
            conn.execute(text(
                "ALTER TABLE projects ADD COLUMN IF NOT EXISTS phase VARCHAR(50) DEFAULT 'Gate B'"
            ))
            conn.commit()
        except Exception:
            conn.rollback()

    db = SessionLocal()
    try:
        if db.query(User).first():
            print("Database already seeded. Skipping.")
            return

        # ── Project Manager User ──────────────────────────────
        pm_user = User(
            email="pm@pmhub.com",
            full_name="Alex Johnson",
            hashed_password=hash_password("password123"),
            role="project_manager",
        )
        db.add(pm_user)
        db.flush()
        print("Created project manager user: pm@pmhub.com")

        # ── Organization & Programme ─────────────────────────
        org = Organization(
            name="UK Power Networks",
            code="UKPN",
            description="UK Power Networks - electricity distribution network operator",
        )
        db.add(org)
        db.flush()
        print(f"Created organization: {org.name} (id={org.id})")

        programme = Programme(
            organization_id=org.id,
            name="Network Upgrade Programme",
            code="NUP",
            description="Infrastructure upgrade and network reinforcement programme",
            status="Active",
        )
        db.add(programme)
        db.flush()
        print(f"Created programme: {programme.name} (id={programme.id})")

        # ── Import Primavera P6 Excel files ───────────────────
        print("Importing Primavera P6 Excel files...")
        for filename in EXCEL_FILES:
            filepath = os.path.join(EXCEL_DIR, filename)
            if os.path.exists(filepath):
                try:
                    import_excel(db, filepath, filename, programme_id=programme.id)
                except Exception as e:
                    print(f"    ERROR importing {filename}: {e}")
            else:
                print(f"    File not found: {filepath}")

        db.commit()
        print("Database seeded successfully!")
    except Exception as e:
        db.rollback()
        print(f"Seeding failed: {e}")
        sys.exit(1)
    finally:
        db.close()


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--update":
        update_excel_data()
    else:
        seed()
